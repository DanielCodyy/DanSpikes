"""
Generate a test Tokyo 2-day trip Excel file for testing the trip upload framework.
Run: python gen_test_trip.py
Output: tokyo2025-test.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT_FILE = os.path.join(os.path.dirname(__file__), 'tokyo2025-test.xlsx')

def style_header(ws, row=1):
    gold = PatternFill("solid", fgColor="1A1A1A")
    hdr_font = Font(bold=True, color="C9A85C", size=10)
    for cell in ws[row]:
        cell.fill = gold
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)

wb = openpyxl.Workbook()

# ──────────────────────────────
# Sheet 1: Info
# ──────────────────────────────
ws_info = wb.active
ws_info.title = 'Info'
ws_info.append(['Key', 'Value'])
style_header(ws_info)
info_rows = [
    ('name_en',    'Tokyo 2025'),
    ('name_zh',    '东京 2025'),
    ('dates_en',   'May 15-16, 2025'),
    ('dates_zh',   '2025年5月15-16日'),
    ('sub_en',     'May 15-16, 2025'),
    ('sub_zh',     '2025年5月15-16日'),
    ('status',     'done'),
    ('cover_url',  'https://picsum.photos/seed/tokyo2025/600/400'),
    ('city_en',    'Tokyo'),
    ('city_zh',    '东京'),
    ('departure_date', '2025-05-15'),
]
for row in info_rows:
    ws_info.append(row)
auto_width(ws_info)

# ──────────────────────────────
# Sheet 2: Days
# ──────────────────────────────
ws_days = wb.create_sheet('Days')
ws_days.append(['date', 'dayName_en', 'dayName_zh', 'area_en', 'area_zh', 'intensity'])
style_header(ws_days)
days = [
    ('2025-05-15', 'Thu · Arrival',   '周四 · 抵达',   'Shinjuku / Harajuku',  '新宿 / 原宿',   'moderate'),
    ('2025-05-16', 'Fri · City Loop', '周五 · 城市环线', 'Asakusa / Shibuya',   '浅草 / 涩谷',   'high'),
]
for row in days:
    ws_days.append(row)
auto_width(ws_days)

# ──────────────────────────────
# Sheet 3: Locations
# ──────────────────────────────
ws_locs = wb.create_sheet('Locations')
headers = [
    'day_date', 'id', 'time',
    'name_en', 'name_zh',
    'lat', 'lng', 'maps_link',
    'special', 'optional',
    'desc_en', 'desc_zh',
    'reservation', 'cost_en', 'cost_zh',
    'booking_url', 'notes_en', 'notes_zh'
]
ws_locs.append(headers)
style_header(ws_locs)

locs = [
    # Day 1
    ('2025-05-15', 'nrt', '08:00', "Narita Airport T1", "成田机场 T1", 35.7719, 140.3929,
     'https://www.google.com/maps/search/?api=1&query=Narita+Airport+Terminal+1',
     'airport', '', "Arrive at NRT", "抵达成田机场", 'fixed', '', '', '', '', ''),
    ('2025-05-15', 'hotel-shinjuku', '11:00', "Hotel Gracery Shinjuku", "新宿格雷斯利酒店", 35.6944, 139.7003,
     'https://www.google.com/maps/search/?api=1&query=Hotel+Gracery+Shinjuku',
     'hotel', '', "Check in or store luggage", "入住或寄存行李", 'fixed', '¥15,000/night', '约¥15,000/晚',
     'https://gracery.com/shinjuku/', 'Godzilla head visible from high floors.', '高层可见哥斯拉雕像。'),
    ('2025-05-15', 'shinjuku-gyoen', '13:00-15:00', "Shinjuku Gyoen", "新宿御苑", 35.6852, 139.7100,
     'https://www.google.com/maps/search/?api=1&query=Shinjuku+Gyoen',
     '', '', "National garden with cherry trees", "国家花园", 'rec', '¥500', '¥500',
     'https://www.env.go.jp/garden/shinjukugyoen/english/', '', ''),
    ('2025-05-15', 'harajuku', '15:30-17:30', "Harajuku / Takeshita Street", "原宿 / 竹下通", 35.6702, 139.7027,
     'https://www.google.com/maps/search/?api=1&query=Takeshita+Street+Harajuku',
     '', '', "Quirky fashion street", "特色时尚街道", 'free', 'Free', '免费', '', '', ''),
    ('2025-05-15', 'meiji-shrine', '17:30-18:30', "Meiji Shrine", "明治神宫", 35.6764, 139.6993,
     'https://www.google.com/maps/search/?api=1&query=Meiji+Shrine+Tokyo',
     '', '', "Forested Shinto shrine", "森林神社", 'free', 'Free', '免费', '', 'Close at sunset.', '日落关闭。'),
    ('2025-05-15', 'shinjuku-night', '19:30', "Shinjuku Night / Kabukicho", "新宿夜 / 歌舞伎町", 35.6952, 139.7038,
     'https://www.google.com/maps/search/?api=1&query=Kabukicho+Shinjuku',
     '', 'true', "Neon-lit entertainment district", "霓虹娱乐街区", 'free', 'Free', '免费', '', 'Stay on main streets at night.', '夜间走大路。'),
    # Day 2
    ('2025-05-16', 'asakusa', '09:00-11:00', "Senso-ji Temple / Asakusa", "浅草寺 / 浅草", 35.7148, 139.7967,
     'https://www.google.com/maps/search/?api=1&query=Sensoji+Temple+Asakusa',
     '', '', "Tokyo's oldest temple", "东京最古老的寺庙", 'free', 'Free', '免费', '', 'Go early to avoid crowds.', '早去人少。'),
    ('2025-05-16', 'akihabara', '11:30-13:00', "Akihabara Electric Town", "秋叶原电器街", 35.7024, 139.7742,
     'https://www.google.com/maps/search/?api=1&query=Akihabara+Tokyo',
     '', 'true', "Electronics and anime district", "电器和动漫街", 'free', 'Free', '免费', '', '', ''),
    ('2025-05-16', 'tsukiji', '13:30-14:30', "Tsukiji Outer Market", "筑地外场市场", 35.6655, 139.7707,
     'https://www.google.com/maps/search/?api=1&query=Tsukiji+Outer+Market+Tokyo',
     '', '', "Fresh seafood and street food", "新鲜海鲜和街头食物", 'free', 'Free', '免费', '', '', ''),
    ('2025-05-16', 'teamlab', '16:00-19:00', "teamLab Borderless", "teamLab 无界", 35.6247, 139.7742,
     'https://www.google.com/maps/search/?api=1&query=teamLab+Borderless+Tokyo',
     '', '', "Immersive digital art museum", "沉浸式数字艺术", 'must', '¥3,200', '¥3,200',
     'https://www.teamlab.art/e/borderless-azabudai/', 'Book well in advance - sells out.', '提前预订，常售罄。'),
    ('2025-05-16', 'shibuya-crossing', '20:00', "Shibuya Scramble Crossing", "涩谷十字路口", 35.6595, 139.7004,
     'https://www.google.com/maps/search/?api=1&query=Shibuya+Scramble+Crossing',
     '', '', "World-famous pedestrian crossing", "世界著名斑马线", 'free', 'Free', '免费', '', '', ''),
]
for row in locs:
    ws_locs.append(list(row))
auto_width(ws_locs)

# ──────────────────────────────
# Sheet 4: Transit
# ──────────────────────────────
ws_transit = wb.create_sheet('Transit')
transit_headers = [
    'day_date', 'from_id', 'to_id', 'mode', 'cost',
    'duration_en', 'duration_zh', 'route_en', 'route_zh',
    'notes_en', 'notes_zh'
]
ws_transit.append(transit_headers)
style_header(ws_transit)

transit = [
    # Day 1
    ('2025-05-15', 'nrt', 'hotel-shinjuku', 'train', '¥3,070',
     '~75 min', '约75分钟',
     'Narita Express (N\'EX) to Shinjuku', '成田特快 (N\'EX) 至新宿',
     'Fastest option from NRT to Shinjuku', '从成田到新宿最快方式'),
    ('2025-05-15', 'hotel-shinjuku', 'shinjuku-gyoen', 'walk', '-',
     '~10 min', '约10分钟', 'Walk from hotel', '从酒店步行', '', ''),
    ('2025-05-15', 'shinjuku-gyoen', 'harajuku', 'metro', '¥178',
     '~5 min', '约5分钟',
     'JR Yamanote Line → Harajuku', 'JR山手线 → 原宿', '', ''),
    ('2025-05-15', 'harajuku', 'meiji-shrine', 'walk', '-',
     '~10 min', '约10分钟', 'Walk to shrine entrance', '步行至神社入口', '', ''),
    ('2025-05-15', 'meiji-shrine', 'shinjuku-night', 'metro', '¥178',
     '~10 min', '约10分钟', 'JR Yamanote to Shinjuku', 'JR山手线至新宿', '', ''),
    # Day 2
    ('2025-05-16', 'hotel-shinjuku', 'asakusa', 'metro', '¥255',
     '~40 min', '约40分钟',
     'Toei Oedo Line → Asakusa', '都营大江户线 → 浅草', '', ''),
    ('2025-05-16', 'asakusa', 'akihabara', 'metro', '¥178',
     '~10 min', '约10分钟', 'Tokyo Metro Ginza Line → Akihabara', '东京地铁银座线 → 秋叶原', '', ''),
    ('2025-05-16', 'akihabara', 'tsukiji', 'metro', '¥178',
     '~15 min', '约15分钟', 'Metro to Tsukiji', '地铁至筑地', '', ''),
    ('2025-05-16', 'tsukiji', 'teamlab', 'metro', '¥204',
     '~25 min', '约25分钟', 'Metro to Azabudai Hills', '地铁至麻布台之丘', '', ''),
    ('2025-05-16', 'teamlab', 'shibuya-crossing', 'metro', '¥178',
     '~15 min', '约15分钟', 'Metro to Shibuya', '地铁至涩谷', '', ''),
]
for row in transit:
    ws_transit.append(list(row))
auto_width(ws_transit)

# ──────────────────────────────
# Sheet 5: Reservations
# ──────────────────────────────
ws_rsv = wb.create_sheet('Reservations')
ws_rsv.append(['priority', 'name_en', 'name_zh', 'when_en', 'when_zh', 'url', 'note_en', 'note_zh'])
style_header(ws_rsv)
rsvs = [
    ('must', 'teamLab Borderless', 'teamLab 无界', 'May 16, 16:00', '5/16 16:00',
     'https://www.teamlab.art/e/borderless-azabudai/', 'Sells out weeks ahead.', '提前数周售罄。'),
    ('rec',  'Shinjuku Gyoen', '新宿御苑', 'May 15, 13:00', '5/15 13:00',
     'https://www.env.go.jp/garden/shinjukugyoen/english/', 'Online entry recommended.', '建议网上购票。'),
    ('opt',  'Narita Express (N\'EX)', '成田特快', 'May 15 morning', '5/15 上午',
     'https://www.jreast.co.jp/e/nex/', 'Book in advance for discounted tickets.', '提前购买优惠票。'),
]
for row in rsvs:
    ws_rsv.append(list(row))
auto_width(ws_rsv)

# ──────────────────────────────
# Sheet 6: Budget
# ──────────────────────────────
ws_bud = wb.create_sheet('Budget')
ws_bud.append(['cat_en', 'cat_zh', 'low', 'high', 'sel', 'note_en', 'note_zh'])
style_header(ws_bud)
budget = [
    ('Hotel, 2 nights', '酒店 2 晚', 140, 220, 180, 'Hotel Gracery Shinjuku', '新宿格雷斯利酒店'),
    ('Narita Express', '成田特快', 28, 32, 30, 'Round trip N\'EX', '成田特快往返'),
    ('Transit (IC card)', '交通 (IC卡)', 20, 35, 28, 'Suica / Pasmo', 'Suica / Pasmo'),
    ('Attractions', '景点门票', 30, 45, 38, 'teamLab + Gyoen', 'teamLab + 新宿御苑'),
    ('Food (2 days)', '餐饮 2 天', 60, 100, 75, 'Mix of ramen, sushi, convenience', '拉面+寿司+便利店'),
    ('Shopping buffer', '购物缓冲', 50, 150, 80, '', ''),
    ('TOTAL', '合计', 328, 582, 431, '', ''),
]
for row in budget:
    ws_bud.append(list(row))
auto_width(ws_bud)

# ──────────────────────────────
# Sheet 7: Tips
# ──────────────────────────────
ws_tips = wb.create_sheet('Tips')
ws_tips.append(['icon', 'title_en', 'title_zh', 'body_en', 'body_zh'])
style_header(ws_tips)
tips = [
    ('🗂️', 'IC Card', 'IC 卡',
     'Get a Suica or Pasmo IC card at any major JR station. Works on almost all trains, buses, and convenience stores.',
     '在各大 JR 站购买 Suica 或 Pasmo，几乎适用于所有电车、公交和便利店。'),
    ('🌡️', 'Weather', '天气',
     'Mid-May in Tokyo: 18-26°C, mostly sunny. Light jacket for evenings.',
     '东京五月中旬 18-26°C，多晴，傍晚可带薄外套。'),
    ('🍜', 'Food', '餐饮',
     'Convenience store onigiri and sandwiches are excellent and cheap. Ramen under ¥1,000 is common.',
     '便利店饭团和三明治极好且便宜。¥1,000 以内的拉面很常见。'),
    ('📱', 'Apps', 'App',
     'Google Maps works well for transit. Hyperdia or Navitime for train schedules.',
     'Google Maps 导航很好用。Hyperdia 或 Navitime 查列车时刻。'),
    ('💴', 'Cash', '现金',
     'Many small restaurants are cash-only. Get yen at the airport or 7-Eleven ATMs.',
     '许多小餐馆只收现金。在机场或 7-11 ATM 取日元。'),
]
for row in tips:
    ws_tips.append(list(row))
auto_width(ws_tips)

# ──────────────────────────────
# Sheet 8: Links
# ──────────────────────────────
ws_links = wb.create_sheet('Links')
ws_links.append(['name_en', 'name_zh', 'url'])
style_header(ws_links)
links = [
    ('JR East (Narita Express)', 'JR东日本 (成田特快)', 'https://www.jreast.co.jp/e/nex/'),
    ('Tokyo Metro Trip Planner', '东京地铁路线', 'https://www.tokyometro.jp/en/tourist/route_map/'),
    ('teamLab Borderless', 'teamLab 无界', 'https://www.teamlab.art/e/borderless-azabudai/'),
    ('Japan Weather (JMA)', '日本气象厅', 'https://www.jma.go.jp/en/week/'),
    ('Hyperdia Train Search', 'Hyperdia 列车搜索', 'https://www.hyperdia.com/en/'),
]
for row in links:
    ws_links.append(list(row))
auto_width(ws_links)

# Save
wb.save(OUT_FILE)
print(f'Saved: {OUT_FILE}')
print(f'Sheets: {wb.sheetnames}')
